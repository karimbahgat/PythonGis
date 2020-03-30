"""
TODO: Switch to a fileformat class that iterates the source data, instead of loading
all into memory... 
"""

# import builtins
import os
import csv
import codecs
import itertools
import warnings

# import fileformat modules
import shapefile as pyshp
import pygeoj

file_extensions = {".shp": "Shapefile",
                   ".json": "GeoJSON",
                   ".geojson": "GeoJSON",
                   ".xls": "Excel 97",
                   ".xlsx": "Excel",
                   ".dta": "Stata",
                   ".csv": "CSV",
                   ".txt": "Text-Delimited",
                   }

def detect_filetype(filepath):
    for ext in file_extensions.keys():
        if filepath.lower().endswith(ext):
            return file_extensions[ext]
    else:
        return None




def from_file(filepath, encoding="utf8", encoding_errors="strict", crs=None, **kwargs):

    # TODO: for geoj and delimited should detect and force consistent field types in similar manner as when saving

    filetype = detect_filetype(filepath)
    
    select = kwargs.get("select")

    def decode(value):
        if isinstance(value, bytes): 
            return value.decode(encoding, errors=encoding_errors)
        else: return value
    
    # shapefile
    if filetype == "Shapefile":
        shapereader = pyshp.Reader(filepath, encoding=encoding, encodingErrors=encoding_errors, **kwargs) # TODO: does pyshp take kwargs?
        
        # load fields, rows, and geometries
        fields = [fieldinfo[0] for fieldinfo in shapereader.fields[1:]]
        rows = ( [value for value in record] for record in shapereader.iterRecords() )
        def getgeoj(obj):
            if obj.shapeTypeName == 'NULL':
                return
            geoj = obj.__geo_interface__
            if hasattr(obj, "bbox"): geoj["bbox"] = list(obj.bbox)
            return geoj
        geometries = (getgeoj(shape) for shape in shapereader.iterShapes())
        rowgeoms = itertools.izip(rows, geometries)
        
        # load projection string from .prj file if exists
        if not crs:
            if os.path.lexists(filepath[:-4] + ".prj"):
                crs = open(filepath[:-4] + ".prj", "r").read()
            else: crs = None

    # geojson file
    elif filetype == "GeoJSON":
        geojfile = pygeoj.load(filepath, **kwargs)

        # load fields, rows, and geometries
        fields = [decode(field) for field in geojfile.common_attributes]
        rows = ([decode(feat.properties[field]) for field in fields] for feat in geojfile)
        geometries = (feat.geometry.__geo_interface__ for feat in geojfile)
        rowgeoms = itertools.izip(rows, geometries)

        # load crs
        if not crs:
            crs = geojfile.crs

    # table files without geometry
    elif filetype in ("Text-Delimited","CSV","Excel 97","Excel","Stata"):

        # txt or csv
        if filetype in ("Text-Delimited","CSV"):
            delimiter = kwargs.get("delimiter")
            fileobj = open(filepath, "rb")
            # auto detect delimiter
            # NOTE: only based on first 10 mb, otherwise gets really slow for large files
            # TODO: run sniffer regardless, and allow sending all kwargs to overwrite
            sniffsize = kwargs.pop('sniffsize', 10)
            dialect = csv.Sniffer().sniff(fileobj.read(1056*sniffsize)) 
            fileobj.seek(0)
            # overwrite with user input
            for k,v in kwargs.items():
                setattr(dialect, k, v)
            # load and parse
            rows = csv.reader(fileobj, dialect)
            def parsestring(string):
                try:
                    val = float(string.replace(",","."))
                    if val.is_integer():
                        val = int(val)
                    return val
                except:
                    if string.upper() == "NULL":
                        return None
                    else:
                        return string.decode(encoding, errors=encoding_errors)
            rows = ([parsestring(cell) for cell in row] for row in rows)
            
            if "skip" in kwargs:
                for _ in range(kwargs["skip"]):
                    next(rows)

            if "last" in kwargs:
                last = kwargs["last"]
                rows = (r for i,r in enumerate(rows) if i <= last)

            fields = next(rows)

        # excel
        elif filetype in ("Excel","Excel 97"):
            if filetype == "Excel 97":
                import xlrd
                wb = xlrd.open_workbook(filepath, encoding_override=encoding, on_demand=True)
                if "sheet" in kwargs:
                    sheet = wb.sheet_by_name(kwargs["sheet"])
                else:
                    sheet = wb.sheet_by_index(0)
                rows = ([cell.value for cell in row] for row in sheet.get_rows())
                
            elif filetype == "Excel":
                import openpyxl as pyxl
                wb = pyxl.load_workbook(filepath, read_only=True)
                if "sheet" in kwargs:
                    sheet = wb[kwargs["sheet"]]
                else:
                    sheet = wb[wb.sheetnames[0]]
                rows = ([cell.value for cell in row] for row in sheet.iter_rows())

            # some excel files may contain junk metadata near top and bottom rows that should be skipped
            # TODO: doesnt actually work for excel, must instead be set with API kw min_row/max_row which
            # ...are wrongly autodetected.
            
            if "skip" in kwargs:
                for _ in range(kwargs["skip"]):
                    next(rows)

            if "last" in kwargs:
                last = kwargs["last"]
                rows = (r for i,r in enumerate(rows) if i <= last)

            fields = next(rows)

        # stata
        elif filetype == "Stata":
            # TODO: how about encoding, manual or pass it on? 
            from .fileformats.stata import StataFile
            dta = StataFile(filepath, encoding=encoding, **kwargs)
            rows = (r for r in dta)
            fields = list(dta.fieldnames)
        
        geokey = kwargs.get("geokey")
        xfield = kwargs.get("xfield")
        yfield = kwargs.get("yfield")
        
        if geokey:
            rowgeoms = ((row,geokey(dict(zip(fields,row)))) for row in rows)
            
        elif xfield and yfield:
            def xygeoj(row):
                rowdict = dict(zip(fields,row))
                x,y = rowdict[xfield],rowdict[yfield]
                try:
                    x,y = float(x),float(y)
                    geoj = {"type":"Point", "coordinates":(x,y)}
                except:
                    try:
                        x,y = float(x.replace(",",".")),float(y.replace(",","."))
                        geoj = {"type":"Point", "coordinates":(x,y)}
                    except:
                        warnings.warn("Could not create point geometry from xfield and yfield values {x} and {y}".format(x=repr(x), y=repr(y)))
                        geoj = None
                return geoj
            rowgeoms = ((row,xygeoj(row)) for row in rows)
            
        else:
            rowgeoms = ((row,None) for row in rows)
            
        crs = None
    
    else:
        raise Exception("Could not create vector data from the given filepath: the filetype extension is either missing or not supported")

    # filter if needed
    if select:
        rowgeoms = ( (row,geom) for row,geom in rowgeoms if select(dict(zip(fields,row))) )

    # load to memory in lists
    rows,geometries = itertools.izip(*rowgeoms)
    rows = list(rows)
    geometries = list(geometries)

    return fields, rows, geometries, crs






